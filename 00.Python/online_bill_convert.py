#! /usr/bin/python3
# _*_ coding:UTF-8 _*_
import os.path
from openpyxl import load_workbook
import csv
import datetime
from queue import Queue

"""
using this python script to convert weixin xlsx file which convert by adobe_acrobat_dc_2022 from pdf files
this file contains at most one year consumption situation, using this to covert xlsx to standard analysis csv file
it also can covert csv data to standard data format for ssj data import to sqlite database
"""

standard_format_data_key = ["交易号", "商家订单号", "交易创建时间", "付款时间", "最近修改时间", "交易对方", "商品名称", "金额", "收/支", "交易状态", "交易方式", "服务费", "备注", "资金状态", "数据来源"]


def covert_weixin_xlsx_bill_to_csv(input_dirs=None):
    if input_dirs is None:
        input_dirs = '../f_input/'
    if not os.path.exists(input_dirs):
        print(f'目录{input_dirs}不存在')
        return
    files = os.listdir(input_dirs)
    weixin_bill_data_list = []
    trade_id_set = set()
    for f in files:
        if f.startswith('微信支付交易明细证明') and f.endswith('.xlsx'):
            xlsx_file = f'{input_dirs}{f}'
            print(f'开始处理-->{xlsx_file}')
            wb = load_workbook(xlsx_file)
            ws = wb.active
            read_data = False
            for r in list(ws.rows):
                if r is not None and r[0] is not None and isinstance(r[0].value, str) and r[0].value.strip().startswith('交易单号'):
                    read_data = True
                    continue
                if r is not None and r[0] is not None and isinstance(r[0].value, str) and r[0].value.strip().startswith('说明'):
                    break
                if read_data:
                    if r[0] is None:
                        continue
                    all_none = True
                    index = 0
                    cell_value_list = []

                    while index < 8 and index < len(r):
                        if r[index] is not None and r[index].value is not None:
                            all_none = False
                        if r[index] is not None and r[index].value is not None:
                            cell_value = r[index].value
                            if isinstance(cell_value, int):
                                cell_value = str(cell_value)
                            if isinstance(cell_value, float):
                                cell_value = str(cell_value)
                            if index == 1:
                                cell_value_list.append(cell_value.replace('\n', ' '))
                            else:
                                cell_value_list.append(cell_value.replace('\n', ''))
                        else:
                            cell_value_list.append(" ")
                        index = index + 1
                    if all_none:
                        # print('have_none')
                        continue

                    if trade_id_set.__contains__(cell_value_list[0]):
                        print(f'该单号已存在{cell_value_list[0]} --> 日期:{cell_value_list[1]}')
                        continue
                    trade_id_set.add(cell_value_list[0])
                    if len(cell_value_list) < 8:
                        print(f"cell_list小于8-->{cell_value_list}")
                        continue
                    weixin_data_tuple = (f'{cell_value_list[0]}\t', f'{cell_value_list[7]}\t', cell_value_list[1], cell_value_list[1], cell_value_list[1], cell_value_list[6], '', cell_value_list[5], cell_value_list[3], '交易成功', cell_value_list[4], '0', cell_value_list[2], '')
                    # print(weixin_data_tuple)
                    weixin_bill_data_list.append(weixin_data_tuple)
            wb.close()
    return weixin_bill_data_list


def convert_weixin_error_xlsx_to_csv(input_dirs=None):
    if input_dirs is None:
        input_dirs = '../f_input/error/'
    if not os.path.exists(input_dirs):
        print(f'目录{input_dirs}不存在')
        return
    files = os.listdir(input_dirs)
    weixin_bill_data_list = []
    trade_id_set = set()
    for f in files:
        if f.startswith('微信支付交易明细证明') and f.endswith('.xlsx'):
            xlsx_file = f'{input_dirs}{f}'
            print(f'开始处理-->{xlsx_file}')
            wb = load_workbook(xlsx_file)
            ws = wb.active
            read_data = False

            row_data_list = list(ws.rows)
            row_index = 0
            while row_index < len(row_data_list):
                r = row_data_list[row_index]
                if r is not None and r[0] is not None and isinstance(r[0].value, str) and r[0].value.strip().startswith('交易单号'):
                    read_data = True
                    row_index = row_index + 1
                    continue
                if r is not None and r[0] is not None and isinstance(r[0].value, str) and r[0].value.strip().startswith('说明'):
                    break
                if read_data:
                    cell_value_list = []
                    if r[3] is not None:
                        flag = covert_type_remove_blank(r[3].value)
                        if flag.__eq__('支出') or flag.__eq__('收入') or flag.__eq__('其他'):
                            for i in range(8):
                                cell_value_list.append(covert_type_remove_blank(r[i].value))

                    row_index = row_index + 1
                    if row_index >= len(row_data_list):
                        break
                    r = row_data_list[row_index]
                    while r[3] is None or r[3].value is None or (isinstance(r[3].value, str) and len(r[3].value) == 0):
                        if row_index >= len(row_data_list):
                            break
                        r = row_data_list[row_index]
                        for i in range(8):
                            if r[i] is not None and r[i].value is not None:
                                cell_value_list[i] = cell_value_list[i] + covert_type_remove_blank(r[i].value)
                        row_index = row_index + 1
                        if row_index >= len(row_data_list):
                            break
                        r = row_data_list[row_index]
                    row_index = row_index - 1
                    if trade_id_set.__contains__(cell_value_list[0]):
                        print(f'该单号已存在{cell_value_list[0]} --> 日期:{cell_value_list[1]}')
                        continue
                    trade_id_set.add(cell_value_list[0])
                    weixin_data_tuple = (f'{cell_value_list[0]}\t', f'{cell_value_list[7]}\t', cell_value_list[1], cell_value_list[1], cell_value_list[1], cell_value_list[6], '', cell_value_list[5], cell_value_list[3], '', cell_value_list[4], '交易成功', cell_value_list[2], '')
                    weixin_bill_data_list.append(weixin_data_tuple)
                row_index = row_index + 1
            wb.close()
    return weixin_bill_data_list
    pass


def covert_type_remove_blank(input_data):
    if input_data is None:
        return input_data
    if isinstance(input_data, datetime.datetime):
        input_data = input_data.strftime("%Y-%m-%d")
    elif isinstance(input_data, datetime.time):
        input_data = " " + input_data.strftime("%H:%M:%S")
    elif not isinstance(input_data, str):
        input_data = str(input_data).strip()
    return input_data


def write_to_standard_weixin_bill_csv(weixin_standard_data_list):
    csv_header = (
    "交易号", "商家订单号", "交易创建时间", "付款时间", "最近修改时间", "交易对方", "商品名称", "金额", "收/支", "交易状态", "交易方式", "服务费", "备注", "资金状态")
    csv_file_name = 'standard_bill_from_weixin.csv'
    if os.path.exists(csv_file_name):
        os.remove(csv_file_name)
    if weixin_standard_data_list is None or len(weixin_standard_data_list) <= 0:
        print('无数据用于写入!!!')
    with open(csv_file_name, 'w', encoding='UTF-8', newline='') as f:
        write = csv.writer(f)
        write.writerow(csv_header)
        for data in weixin_standard_data_list:
            write.writerow(data)


def print_cell_in_row(row):
    for c in row:
        if c.value is None:
            continue
        cell_value = c.value
        if isinstance(c.value, str):
            cell_value = c.value.replace('\n', '')
        print(f"{cell_value}", end=', ')


def save_csv_file():


    pass


def parseTradeID():
    input_file_name = '../f_input/微信交易单号.txt'
    trade_id_list = []
    data_queue = Queue()
    with open(input_file_name, mode="r+", encoding="UTF-8") as f:
        data = f.read()
        data = data.replace('\n', '')
        for i in range(0, len(data)):
            data_char = data[i:i + 1]
            if data_char == '"':
                if not data_queue.empty():
                    tp_str = ' '
                    while not data_queue.empty():
                        tp_str = tp_str + data_queue.get()
                    if len(tp_str) > 0:
                        trade_id_list.append(tp_str.strip())
            else:
                data_queue.put(data_char)
    # print(trade_id_list)
    print(len(trade_id_list))
    data_set = set()
    for i in trade_id_list:
        if data_set.__contains__(i):
            print(i)
        data_set.add(i)
    input_standard_file = '../f_input/standard_bill_from_weixin.csv'
    with open(input_standard_file, 'r', encoding="UTF-8") as f:
        reader = csv.reader(f)
        for row in reader:
            if row[0].strip().__eq__('交易号'):
                continue
            data_set.discard(row[0].strip())

    print('----------------------------------------------------')
    print(data_set)
    print(len(data_set))


def convert_alipay_bill_to_standard_data_format():
    """
    convert alipay csv data to standard format ssj data to import sqlite database
    :return:
    """
    input_dir = '../f_input/'
    files = os.listdir(input_dir)
    standard_data_list = []
    for f in files:
        if f.startswith("alipay_record") and f.endswith(".csv"):
            with open(f'{input_dir}{f}', 'r') as f:
                reader = csv.reader(f)
                data_header = []
                data_read_flag = False
                for row in reader:
                    if row[0].strip().__eq__('交易号'):
                        data_read_flag = True
                        data_header = row
                        continue
                    if data_read_flag:
                        if len(row) < 5:
                            continue
                        standard_data = dict.fromkeys(standard_format_data_key)
                        column_index = 0
                        while column_index < max(len(data_header), len(standard_format_data_key)):
                            standard_data[data_header[column_index].strip()] = row[column_index].replace('\t', '').strip()
                            column_index = column_index + 1
                        standard_data['金额'] = standard_data['金额（元）']
                        standard_data['服务费'] = standard_data['服务费（元）']
                        standard_data.pop('金额（元）')
                        standard_data.pop('服务费（元）')
                        standard_data.pop('交易来源地')
                        standard_data.pop('类型')
                        standard_data.pop('成功退款（元）')
                        standard_data['交易方式'] = '支付宝'
                        if standard_data['资金状态'] is None:
                            standard_data['资金状态'] = ' '
                        if standard_data['备注'] is None:
                            standard_data['备注'] = ' '
                        standard_data['备注'] = f'{standard_data["备注"]}\n数据来源:支付宝'
                        standard_data['数据来源'] = 'from_alipay_bill'
                        standard_data_list.append(standard_data)
    print(f'支付宝待插入数据{len(standard_data_list)}条')
    return standard_data_list


def convert_weixin_bill_to_standard_data_format():
    """
    convert
    :return:
    """
    input_dir = '../f_input/'
    files = os.listdir(input_dir)
    standard_data_list = []
    for f in files:
        if f.startswith("微信支付账单") and f.endswith(".csv"):
            with open(f'{input_dir}{f}', 'r', encoding="UTF-8") as f:
                reader = csv.reader(f)
                data_header = []
                data_read_flag = False
                for row in reader:
                    if row[0].strip().__eq__('交易时间'):
                        data_read_flag = True
                        data_header = row
                        continue
                    if data_read_flag:
                        if len(row) < 5:
                            continue
                        standard_data = dict.fromkeys(standard_format_data_key)
                        column_index = 0
                        while column_index < min(len(data_header), len(standard_format_data_key)):
                            standard_data[data_header[column_index].strip()] = row[column_index].replace('\t', '').strip()
                            column_index = column_index + 1
                        standard_data['金额'] = standard_data['金额(元)'].replace('¥', '')
                        standard_data['交易创建时间'] = standard_data['交易时间']
                        standard_data['付款时间'] = standard_data['交易时间']
                        standard_data['最近修改时间'] = standard_data['交易时间']
                        standard_data['商品名称'] = standard_data['商品']
                        standard_data['交易号'] = standard_data['交易单号']
                        standard_data['商家订单号'] = standard_data['商户单号']
                        standard_data['交易状态'] = '交易成功'
                        standard_data['服务费'] = '0'

                        pay_channel = standard_data['支付方式']
                        if pay_channel.__contains__('银行'):
                            standard_data['交易方式'] = '银行卡'
                        else:
                            standard_data['交易方式'] = '微信钱包'

                        if standard_data['资金状态'] is None:
                            standard_data['资金状态'] = ' '
                        if standard_data['备注'] is None:
                            standard_data['备注'] = ' '
                        standard_data['备注'] = standard_data['备注'].replace('/', '')
                        standard_data['备注'] = f'{standard_data["备注"]}\n{standard_data["当前状态"]}\n数据来源:微信短账单'
                        standard_data['数据来源'] = 'from_weixin_bill'
                        standard_data.pop('交易时间')
                        standard_data.pop('商品')
                        standard_data.pop('金额(元)')
                        standard_data.pop('交易类型')
                        standard_data.pop('交易单号')
                        standard_data.pop('商户单号')
                        standard_data.pop('支付方式')
                        standard_data.pop('当前状态')
                        standard_data_list.append(standard_data)
    print(f'微信待插入数据{len(standard_data_list)}条')
    # print(standard_data_list)
    return standard_data_list


def convert_weixin_long_bill_to_standard_data_format():
    input_dir = '../f_input/'
    files = os.listdir(input_dir)
    standard_data_list = []
    for f in files:
        if f.startswith("standard_bill_from_weixin") and f.endswith(".csv"):
            with open(f'{input_dir}{f}', 'r', encoding='UTF-8') as f:
                reader = csv.reader(f)
                data_header = []
                is_first = True
                for row in reader:
                    if is_first:
                        data_header = row
                        is_first = False
                        continue
                    standard_data = dict.fromkeys(standard_format_data_key)
                    column_index = 0
                    while column_index < len(data_header):
                        standard_data[data_header[column_index]] = row[column_index].replace('\t', '')
                        column_index = column_index + 1
                    if standard_data['备注'] is None:
                        standard_data['备注'] = ' '
                    standard_data['备注'] = f'{standard_data["备注"]}\n数据来源:微信支付'
                    if standard_data['交易方式'] is not None:
                        if standard_data['交易方式'].strip().__contains__('零钱'):
                            standard_data['交易方式'] = '微信钱包'
                        elif standard_data['交易方式'].strip().__contains__('银行'):
                            standard_data['交易方式'] = '银行卡'
                        elif standard_data['交易方式'].strip().__eq__('/'):
                            standard_data['交易方式'] = '微信钱包'
                    else:
                        standard_data['交易方式'] = '微信钱包'
                    standard_data['数据来源'] = 'from_weixin_bill'
                    standard_data_list.append(standard_data)
    print(f'微信待插入数据{len(standard_data_list)}条')
    return standard_data_list
    pass


if __name__ == '__main__':
    weixin_data_normal = covert_weixin_xlsx_bill_to_csv()
    print(len(weixin_data_normal))
    weixin_data_error = convert_weixin_error_xlsx_to_csv()
    print(len(weixin_data_error))
    weixin_standard_data = weixin_data_normal + weixin_data_error
    weixin_trade_id_set = set()
    for i in weixin_standard_data:
        if weixin_trade_id_set.__contains__(i[0]):
            print(f'重复数据-->{i}')
    print(len(weixin_standard_data))
    # write_to_standard_weixin_bill_csv(weixin_standard_data)
    pass